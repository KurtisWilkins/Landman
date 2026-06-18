"""Tabular parsing for CSV and Excel (design doc §5.2).

Both return ``(headers, rows)`` where rows are dicts keyed by header. Excel uses openpyxl;
CSV uses the stdlib. Values are kept as strings; the normalizers coerce numbers (so a stray
non-numeric cell never fails the whole ingest — greedy ingest).
"""

from __future__ import annotations

import csv
import io
from collections.abc import Sequence

Row = dict[str, str]


def _rows_from_records(headers: Sequence[str], records: Sequence[Sequence[object]]) -> list[Row]:
    rows: list[Row] = []
    for rec in records:
        row = {
            headers[i]: ("" if rec[i] is None else str(rec[i]))
            for i in range(min(len(headers), len(rec)))
        }
        if any(v.strip() for v in row.values()):
            rows.append(row)
    return rows


def parse_csv(data: bytes) -> tuple[list[str], list[Row]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = list(csv.reader(io.StringIO(text)))
    if not reader:
        return [], []
    headers = [h.strip() for h in reader[0]]
    return headers, _rows_from_records(headers, reader[1:])


def parse_xlsx(data: bytes) -> tuple[list[str], list[Row]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return [], []
    records = list(ws.iter_rows(values_only=True))
    if not records:
        return [], []
    headers = [("" if c is None else str(c)).strip() for c in records[0]]
    return headers, _rows_from_records(headers, records[1:])


def parse_tabular(data: bytes, content_type: str, filename: str) -> tuple[list[str], list[Row]]:
    """Dispatch to the right parser by content type / extension."""
    name = filename.lower()
    if name.endswith((".xlsx", ".xlsm")) or "spreadsheet" in content_type:
        return parse_xlsx(data)
    return parse_csv(data)


def parse_csv_matrix(data: bytes) -> list[list[object]]:
    text = data.decode("utf-8-sig", errors="replace")
    return [list(rec) for rec in csv.reader(io.StringIO(text))]


def parse_xlsx_matrix(data: bytes) -> list[list[object]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    return [list(rec) for rec in ws.iter_rows(values_only=True)]


def parse_matrix(data: bytes, content_type: str, filename: str) -> list[list[object]]:
    """Raw cell matrix (no header assumption) — for layouts whose header isn't row 0,
    e.g. the QuickBooks 'N Month Recap' P&L."""
    name = filename.lower()
    if name.endswith((".xlsx", ".xlsm")) or "spreadsheet" in content_type:
        return parse_xlsx_matrix(data)
    return parse_csv_matrix(data)
